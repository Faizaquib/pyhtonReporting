#!/usr/bin/env python
# coding: utf-8

# In[4]:


from datetime import datetime
import openpyxl
import fnmatch
import os
import tkinter as tk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import numpy as np

LARGE_FONT= ("Verdana", 12)


class SeaofBTCapp(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)

        container.pack(side="top", fill="both", expand = True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (StartPage, January,Feburary,March,April,May,June,July,August,September,October,November,December):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)

    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()

        
class StartPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        label = tk.Label(self, text="Hey There!", font=LARGE_FONT)
        label.pack(pady=50,padx=50)

        button = tk.Button(self, text="January,2020",
                            command=lambda: controller.show_frame(January))
        button.pack(side='left',pady=10,padx=10)

        button2 = tk.Button(self, text="Feburary,2020",
                            command=lambda: controller.show_frame(Feburary))
        button2.pack(side='left',pady=10,padx=10)
        
        button = tk.Button(self, text="March,2020",
                            command=lambda: controller.show_frame(March))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="April,2020",
                            command=lambda: controller.show_frame(April))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="May,2020",
                            command=lambda: controller.show_frame(May))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="June,2020",
                            command=lambda: controller.show_frame(June))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="July,2020",
                            command=lambda: controller.show_frame(July))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="August,2020",
                            command=lambda: controller.show_frame(August))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="September,2020",
                            command=lambda: controller.show_frame(September))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="October,2020",
                            command=lambda: controller.show_frame(October))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="November,2020",
                            command=lambda: controller.show_frame(November))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="December,2020",
                            command=lambda: controller.show_frame(December))
        button.pack(side='left',pady=10,padx=10)

class January(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        
        #print('loading modules....1/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\january.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in January,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        

class Feburary(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....2/12')
        wb = openpyxl.load_workbook(r'\P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\feburary.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                 name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in Feburary,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        

class March(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....3/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\march.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in March,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)


class April(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....4/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\april.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in April,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class May(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....5/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\may.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in May,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class June(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....6/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\june.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in June,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class July(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....7/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\july.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in July,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class August(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....8/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\august.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in August,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class September(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....9/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\september.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in September,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class October(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....10/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\october.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in October,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class November(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....11/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\november.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in November,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class December(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        #print('loading modules....12/12')
        wb = openpyxl.load_workbook(r'P:\Gateshead Ops\NS36 Workaround\CIM\Team Members\Faiz\Reporting\genesisOne\december.xlsx')
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            if (sheet.cell(row=i,column=2).value > 0):
                name.append(sheet.cell(row=i,column=1).value)
                iteration.append(sheet.cell(row=i,column=2).value)
                timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in December,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name,rotation=30)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.30, v, iteration[i],fontsize=7)
            for i, v in enumerate(timeTaken):
                a.text(i+.00, v, timeTaken[i],fontsize=7)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        




                    

app = SeaofBTCapp()
app.geometry('1400x850')
app.mainloop()
#app.title("Genesis's Work!")


# In[ ]:




