#!/usr/bin/env python
# coding: utf-8

# In[8]:


from datetime import datetime
import openpyxl
import fnmatch
import os
import tkinter as tk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import numpy as np
import time

def countDailyFile(day,processName,Month):
    if processName == 'RTB':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\InstallsRTBMultipleTerminal\InstallsRTBMultipleTerminalFilesArchive'
    elif processName == 'MBD':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\MBD\MBDFilesArchive'
    elif processName == 'ManualPayments':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\ManualPayments\ManualPaymentsFilesArchive'
    elif processName =='AutoOnboarding':
        path = r'\\tsclient\P\Applications_XML\AutoCustomerOnboardingFilesArchive'
    elif processName == 'BankAddressChange':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\BankAddress\BankAddressChangeFilesArchive'
    elif processName == 'ADDACS-Code0':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\ADDACS\0\0FilesArchive'
    elif processName == 'ADDACS-Code1':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\ADDACS\1\1FilesArchive'
    elif processName == 'ADDACS-CodeAuddis':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\ADDACS\Auddis\AuddisFilesArchive'
    elif processName == 'ADDACS-CodeCORPORATE':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\ADDACS\CORPORATE\CORPORATEFilesArchive'
    elif processName == 'Alerts':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\Alerts\AlertsFilesArchive'
    elif processName == 'AutoOutletCloning':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\AutoOutletCloning\AutoOutletCloningFilesArchive'
    elif processName == 'RateAmendments':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\RateAmendments\RateAmendmentsFilesArchive'
    elif processName == 'RemovalsRTB':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\RTB - Removal\RemovalsRTBFilesArchive'
    elif processName == 'SRClosures':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\SRClosures\SRClosuresFilesArchive'
    elif processName == 'TallyRollsCharges':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\Tally Roll\TallyRollsChargeFilesArchive'
    elif processName == 'WAF':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\WAF\WAFFilesArchive'    
    
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
    
    if (int(day)<10):
        day = str(0)+str(day)
     
    if (int(month)<10):
        month = str(0)+str(month)
    
    
    listOfFiles = fnmatch.filter(os.listdir(path),"*.txt")
    count = 0
    time = 0.00
    for file in listOfFiles:
        currentDate = str(day)+'-'+str(month)+'-'+str(datetime.now().year)
        fileCreationDate = datetime.date(datetime.fromtimestamp(os.stat(path+"\\"+file).st_mtime)).strftime("%d-%m-%Y")
        #iterator = iterator + 1 
        if currentDate == fileCreationDate :
            count = count + 1
            with open(path+'\\'+file) as f:
                for line in f:
                    if 'Total Time Taken' in line:
                        a = (line.split())
                        minutes = a[5].replace(',','')
                        time = time + float(minutes)
    return count,time


    

def updateExcelTwo(processName,Month):
    wb = openpyxl.load_workbook(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisTwo\\'+Month+'\\'+processName+'.xlsx')
    sheet = wb["Sheet1"] 
    lastDay = str(sheet.cell(row=sheet.max_row,column=3).value)
    currentDay = str(datetime.now().day)
    currentMonth = datetime.now().month
    lastRow = sheet.max_row
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
        
        
    
    if(int(month) < int(currentMonth)):
        if(int(lastDay)<32):
            for i in range(int(lastDay),32):
                sheet.cell(row=lastRow,column=1).value = datetime.now().year
                sheet.cell(row=lastRow,column=2).value = month
                sheet.cell(row=lastRow,column=3).value = i
                count  , time = countDailyFile(i,processName,Month)
                sheet.cell(row=lastRow,column=4).value = count
                sheet.cell(row=lastRow,column=5).value = round(time/60,2)
                wb.save(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisTwo\\'+Month+'\\'+processName+'.xlsx')
                lastRow = lastRow + 1
    elif(int(month) == int(currentMonth)):
        if(int(lastDay)<int(currentDay)):
            for i in range(int(lastDay),int(currentDay)+1):
                sheet.cell(row=lastRow,column=1).value = datetime.now().year
                sheet.cell(row=lastRow,column=2).value = month
                sheet.cell(row=lastRow,column=3).value = i
                count , time = countDailyFile(i,processName,Month)
                sheet.cell(row=lastRow,column=4).value = count
                sheet.cell(row=lastRow,column=5).value = round(time/60,2)
                wb.save(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisTwo\\'+Month+'\\'+processName+'.xlsx')
                lastRow = lastRow + 1
        elif(int(lastDay) == int(currentDay)):
            sheet.cell(row=lastRow,column=1).value = datetime.now().year
            sheet.cell(row=lastRow,column=2).value = month
            sheet.cell(row=lastRow,column=3).value = currentDay
            count , time = countDailyFile(currentDay,processName,Month)
            sheet.cell(row=lastRow,column=4).value = count
            sheet.cell(row=lastRow,column=5).value = round(time/60,2)
            wb.save(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisTwo\\'+Month+'\\'+processName+'.xlsx')
    if (processName=='TallyRollsCharges'):
        print(str(month)+'/12...Modules Loaded')
    


def totalIteration(processName,Month):
    updateExcelTwo(processName,Month)
    totalIteration = 0
    wb = openpyxl.load_workbook(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisTwo\\'+Month+'\\'+processName+'.xlsx')
    sheet = wb["Sheet1"]
    for i in range(2,sheet.max_row+1):
        totalIteration = totalIteration + sheet.cell(row=i,column=4).value
    return totalIteration

def totalTime(processName,Month):
    #pdateExcelTwo(processName,Month)
    totalTime = 0
    wb = openpyxl.load_workbook(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisTwo\\'+Month+'\\'+processName+'.xlsx')
    sheet = wb.get_sheet_by_name("Sheet1")
    for i in range(2,sheet.max_row+1):
        val = sheet.cell(row=i,column=5).value
        totalTime = totalTime + int(0 if val is None else val)
    return totalTime



def updateExcelOne(processName,Month):
    wb = openpyxl.load_workbook(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisOne\\'+Month+'.xlsx')
    sheet = wb["Sheet1"]
    for i in range(2,sheet.max_row+1):
        if(sheet.cell(row=i,column=1).value == processName):
            sheet.cell(row=i,column=2).value = totalIteration(processName,Month)
            sheet.cell(row=i,column=3).value = round(totalTime(processName,Month),2)
            wb.save(r'\\tsclient\\P\\Gateshead Ops\\NS36 Workaround\\CIM\\Team Members\\Faiz\\Reporting\\genesisOne\\'+Month+".xlsx")
            #print('Excel Two updated!')
            #print(Month)
                    
#updateExcelOne('RTB','december')

def caller():
    
    start = time.process_time()
    #print(start)
    currentMonth = int(datetime.now().month)
    i = currentMonth
    if i ==1:
        month = 'january'
    elif i == 2:
        month = 'feburary'
    elif i == 3:
        month = 'march'
    elif i == 4:
        month = 'april'
    elif i == 5:
        month = 'may'
    elif i == 6:
        month = 'june'
    elif i == 7:
        month = 'july'
    elif i == 8:
        month = 'august'
    elif i == 9:
        month = 'september'
    elif i == 10:
        month = 'october'
    elif i == 11:
        month = 'november'   
    elif i == 12:
        month = 'december'
    #print(currentMonth)
    print(month)            
    updateExcelOne('RTB',month)
    print("Completed Processes :")
    print("RTB")
    updateExcelOne('MBD',month)
    print("MBD")
    updateExcelOne('ManualPayments',month)
    print("Manula Payments ")
    updateExcelOne('BankAddressChange',month)
    print("bank Address Change")
    updateExcelOne('AutoOnboarding',month)
    print("autocustomeronbaording")               
    updateExcelOne('ADDACS-Code0',month)
    print("code0")               
    updateExcelOne('ADDACS-Code1',month)
    print("code1")              
    updateExcelOne('ADDACS-CodeAuddis',month)
    print("codeauddis")
    updateExcelOne('ADDACS-CodeCORPORATE',month)
    print("codeCorporate")
    updateExcelOne('Alerts',month)
    print("alerts")
    updateExcelOne('WAF',month)
    print("WAF")
    updateExcelOne('AutoOutletCloning',month)
    print("AutoOutletCloning")
    updateExcelOne('RateAmendments',month)
    print("rateamendments")
    updateExcelOne('RemovalsRTB',month)
    print("removalsRTB")
    updateExcelOne('SRClosures',month)
    print("SRClousre")
    updateExcelOne('TallyRollsCharges',month)
    print("TallyRollsCharges")
    #print(time.process_time() - start)
caller()


# In[ ]:




