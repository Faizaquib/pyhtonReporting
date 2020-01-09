#!/usr/bin/env python
# coding: utf-8

# In[34]:


from datetime import datetime
import openpyxl
import fnmatch
import os
import tkinter as tk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import numpy as np



def countDailyFileExcel(day,processName,Month):
    if processName == 'RTB':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\InstallsRTBMultipleTerminal\\InstallsRTBMultipleTerminalFilesArchive'
    elif processName == 'MBD':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\MBD\\MBDFilesArchive'
    elif processName == 'ManualPayments':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\ManualPayments\\ManualPaymentsFilesArchive'
    elif processName =='AutoOnboarding':
        path = 'P:\\Applications_XML\\AutoCustomerOnboardingFilesArchive'
    elif processName == 'BankAddressChange':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\BankAddress\\BankAddressChangeFilesArchive'
    
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
    
    if (int(day)<10):
        day = str(0)+str(day)
     
    if (int(month)<10):
        month = str(0)+str(month)
    
    
    listOfFiles = fnmatch.filter(os.listdir(path),"*.xlsx")
    count = 0
    for file in listOfFiles:
        currentDate = str(day)+'-'+str(month)+'-'+str(datetime.now().year)
        fileCreationDate = datetime.date(datetime.fromtimestamp(os.stat(path+"\\"+file).st_mtime)).strftime("%d-%m-%Y")
        if currentDate == fileCreationDate :
            wb = openpyxl.load_workbook(path+"\\"+file)
            sheet = wb.get_sheet_by_name("Sheet1")
            count = count + sheet.max_row 
    return count

def calculateTimeExcel(day,processName,Month):
    if processName == 'RTB':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\InstallsRTBMultipleTerminal\\InstallsRTBMultipleTerminalFilesArchive'
    elif processName == 'MBD':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\MBD\\MBDFilesArchive'
    elif processName == 'ManualPayments':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\ManualPayments\\ManualPaymentsFilesArchive'
    elif processName =='AutoOnboarding':
        path = 'P:\\Applications_XML\\AutoCustomerOnboardingFilesArchive'
    elif processName == 'BankAddressChange':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\BankAddress\\BankAddressChangeFilesArchive'
    
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
    
    if (int(day)<10):
        day = str(0)+str(day)
     
    if (int(month)<10):
        month = str(0)+str(month)
    
    listOfFiles = fnmatch.filter(os.listdir(path),"*.xlsx")
    array = []
    count = 0
    time = 0.00
    for file in listOfFiles:
        currentDate = str(day)+'-'+str(month)+'-'+str(datetime.now().year)
        fileCreationDate = datetime.date(datetime.fromtimestamp(os.stat(path+"\\"+file).st_mtime)).strftime("%d-%m-%Y")
        if currentDate == fileCreationDate :
            array.append(file)
    for fileName in range(0,len(array)):
        wb = openpyxl.load_workbook(path+"\\"+array[fileName])
        sheet = wb.get_sheet_by_name("Sheet1")
        count = sheet.max_row
        totalTime = []
        for i in range(1,count+1):
            totalTime.append(sheet.cell(row=i,column=sheet.max_column).value)
        time = time + float( sum(filter(lambda i: isinstance(i, int), totalTime)))
        time = time + sum(filter(lambda i: isinstance(i, float), totalTime))
    return time

def countDailyFile(day,processName,Month):
    if processName == 'RTB':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\InstallsRTBMultipleTerminal\\InstallsRTBMultipleTerminalFilesArchive'
    elif processName == 'MBD':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\MBD\\MBDFilesArchive'
    elif processName == 'ManualPayments':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\ManualPayments\\ManualPaymentsFilesArchive'
    elif processName =='AutoOnboarding':
        path = 'P:\\Applications_XML\\AutoCustomerOnboardingFilesArchive'
    elif processName == 'BankAddressChange':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\BankAddress\\BankAddressChangeFilesArchive'
    
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
    
    if (int(day)<10):
        day = str(0)+str(day)
     
    if (int(month)<10):
        month = str(0)+str(month)
    
    
    listOfFiles = fnmatch.filter(os.listdir(path),"*.txt")
    count = 0
    for file in listOfFiles:
        currentDate = str(day)+'-'+str(month)+'-'+str(datetime.now().year)
        fileCreationDate = datetime.date(datetime.fromtimestamp(os.stat(path+"\\"+file).st_mtime)).strftime("%d-%m-%Y")
        #iterator = iterator + 1 
        if currentDate == fileCreationDate : 
                count = count + 1 
    return count

def calculateTime(day,processName,Month):
    if processName == 'RTB':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\InstallsRTBMultipleTerminal\\InstallsRTBMultipleTerminalFilesArchive'
    elif processName == 'MBD':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\MBD\\MBDFilesArchive'
    elif processName == 'ManualPayments':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\ManualPayments\\ManualPaymentsFilesArchive'
    elif processName =='AutoOnboarding':
        path = 'P:\\Applications_XML\\AutoCustomerOnboardingFilesArchive'
    elif processName == 'BankAddressChange':
        path = 'P:\\Gateshead Ops\\NS36 Workaround\\CIM\\BankAddress\\BankAddressChangeFilesArchive'
    
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
    
    if (int(day)<10):
        day = str(0)+str(day)
     
    if (int(month)<10):
        month = str(0)+str(month)
    
    listOfFiles = fnmatch.filter(os.listdir(path),"*.txt")
    array = []
    time = 0.00
    for file in listOfFiles:
        currentDate = str(day)+'-'+str(month)+'-'+str(datetime.now().year)
        fileCreationDate = datetime.date(datetime.fromtimestamp(os.stat(path+"\\"+file).st_mtime)).strftime("%d-%m-%Y")
        #iterator = iterator + 1 
        if currentDate == fileCreationDate : 
                #count = count + 1
                array.append(file)
    
    for fileName in range(0,len(array)):
        with open(path+'\\'+array[fileName]) as f:
            for line in f:
                 if 'Total Time Taken' in line:
                        a = (line.split())
                        time = time + float(a[5])
    return time
    


def updateExcelTwo(processName,Month):
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    sheet = wb.get_sheet_by_name("Sheet1") 
    lastDay = str(sheet.cell(row=sheet.max_row,column=3).value)
    currentDay = str(datetime.now().day)
    currentMonth = datetime.now().month
    lastRow = sheet.max_row
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
        
        
    if processName in ['RTB','ManualPayments','MBD']:
        if(month < currentMonth):
            if(lastDay<currentDay):
                for i in range(int(lastDay)+1,32):
                    lastRow = lastRow + 1
                    sheet.cell(row=lastRow,column=1).value = datetime.now().year
                    sheet.cell(row=lastRow,column=2).value = month
                    sheet.cell(row=lastRow,column=3).value = i
                    sheet.cell(row=lastRow,column=4).value = countDailyFileExcel(i,processName,Month)
                    sheet.cell(row=lastRow,column=5).value = round(calculateTimeExcel(i,processName,Month)/60,2)
                    wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
        elif(month == currentMonth):
            if(lastDay<currentDay):
                for i in range(int(lastDay)+1,int(currentDay)+1):
                    lastRow = lastRow + 1
                    sheet.cell(row=lastRow,column=1).value = datetime.now().year
                    sheet.cell(row=lastRow,column=2).value = month
                    sheet.cell(row=lastRow,column=3).value = i
                    sheet.cell(row=lastRow,column=4).value = countDailyFileExcel(i,processName,Month)
                    sheet.cell(row=lastRow,column=5).value = round(calculateTimeExcel(i,processName,Month)/60,2)
                    wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
            elif(lastDay==currentDay):
                sheet.cell(row=lastRow,column=1).value = datetime.now().year
                sheet.cell(row=lastRow,column=2).value = month
                sheet.cell(row=lastRow,column=3).value = currentDay
                sheet.cell(row=lastRow,column=4).value = countDailyFileExcel(currentDay,processName,Month)
                sheet.cell(row=lastRow,column=5).value = round(calculateTimeExcel(currentDay,processName,Month)/60,2)
                wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
        
    else :
        if(month < currentMonth):
            if(lastDay<currentDay):
                for i in range(int(lastDay)+1,32):
                    lastRow = lastRow + 1
                    sheet.cell(row=lastRow,column=1).value = datetime.now().year
                    sheet.cell(row=lastRow,column=2).value = month
                    sheet.cell(row=lastRow,column=3).value = i
                    sheet.cell(row=lastRow,column=4).value = countDailyFile(i,processName,Month)
                    sheet.cell(row=lastRow,column=5).value = round(calculateTime(i,processName,Month)/60,2)
                    wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
        elif(month == currentMonth):
            if(lastDay<currentDay):
                for i in range(int(lastDay)+1,int(currentDay)+1):
                    lastRow = lastRow + 1
                    sheet.cell(row=lastRow,column=1).value = datetime.now().year
                    sheet.cell(row=lastRow,column=2).value = month
                    sheet.cell(row=lastRow,column=3).value = i
                    sheet.cell(row=lastRow,column=4).value = countDailyFile(i,processName,Month)
                    sheet.cell(row=lastRow,column=5).value = round(calculateTime(i,processName,Month)/60,2)
                    wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
            elif(lastDay==currentDay):
                sheet.cell(row=lastRow,column=1).value = datetime.now().year
                sheet.cell(row=lastRow,column=2).value = month
                sheet.cell(row=lastRow,column=3).value = currentDay
                sheet.cell(row=lastRow,column=4).value = countDailyFile(currentDay,processName,Month)
                sheet.cell(row=lastRow,column=5).value = round(calculateTime(currentDay,processName,Month)/60,2)
                wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    
    


def totalIteration(processName,Month):
    updateExcelTwo(processName,Month)
    totalIteration = 0
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    for i in range(2,sheet.max_row+1):
        totalIteration = totalIteration + sheet.cell(row=i,column=4).value
    return totalIteration

def totalTime(processName,Month):
    #pdateExcelTwo(processName,Month)
    totalTime = 0
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    for i in range(2,sheet.max_row+1):
        totalTime = totalTime + sheet.cell(row=i,column=5).value
    return totalTime



def updateExcelOne(processName,Month):
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\"+Month+".xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    for i in range(2,sheet.max_row+1):
        if(sheet.cell(row=i,column=1).value == processName):
            sheet.cell(row=i,column=2).value = totalIteration(processName,Month)
            sheet.cell(row=i,column=3).value = totalTime(processName,Month)
            wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\"+Month+".xlsx")
            #print('Excel Two updated!')
            #print(Month)
                    
#updateExcelOne('RTB','december')

LARGE_FONT= ("Verdana", 12)


class SeaofBTCapp(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)

        container.pack(side="top", fill="both", expand = True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (StartPage, January,Feburary,March,April,May,June,July,August,September,October,November,December):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)

    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()

        
class StartPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        label = tk.Label(self, text="Hey There!", font=LARGE_FONT)
        label.pack(pady=50,padx=50)

        button = tk.Button(self, text="January,2020",
                            command=lambda: controller.show_frame(January))
        button.pack(side='left',pady=10,padx=10)

        button2 = tk.Button(self, text="Feburary,2020",
                            command=lambda: controller.show_frame(Feburary))
        button2.pack(side='left',pady=10,padx=10)
        
        button = tk.Button(self, text="March,2020",
                            command=lambda: controller.show_frame(March))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="April,2020",
                            command=lambda: controller.show_frame(April))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="May,2020",
                            command=lambda: controller.show_frame(May))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="June,2020",
                            command=lambda: controller.show_frame(June))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="July,2020",
                            command=lambda: controller.show_frame(July))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="August,2020",
                            command=lambda: controller.show_frame(August))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="September,2020",
                            command=lambda: controller.show_frame(September))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="October,2020",
                            command=lambda: controller.show_frame(October))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="November,2020",
                            command=lambda: controller.show_frame(November))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="December,2020",
                            command=lambda: controller.show_frame(December))
        button.pack(side='left',pady=10,padx=10)

class January(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','january')
        updateExcelOne('MBD','january')
        updateExcelOne('ManualPayments','january')
        updateExcelOne('BankAddressChange','january')
        updateExcelOne('AutoOnboarding','january')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\january.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
            timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in January,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        

class Feburary(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','feburary')
        updateExcelOne('MBD','feburary')
        updateExcelOne('ManualPayments','feburary')
        updateExcelOne('BankAddressChange','feburary')
        updateExcelOne('AutoOnboarding','feburary')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\feburary.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in Feburary,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        

class March(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','march')
        updateExcelOne('MBD','march')
        updateExcelOne('ManualPayments','march')
        updateExcelOne('BankAddressChange','march')
        updateExcelOne('AutoOnboarding','march')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\march.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in March,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)


class April(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','april')
        updateExcelOne('MBD','april')
        updateExcelOne('ManualPayments','april')
        updateExcelOne('BankAddressChange','april')
        updateExcelOne('AutoOnboarding','april')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\april.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in April,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class May(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','may')
        updateExcelOne('MBD','may')
        updateExcelOne('ManualPayments','may')
        updateExcelOne('BankAddressChange','may')
        updateExcelOne('AutoOnboarding','may')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\may.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in May,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class June(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','june')
        updateExcelOne('MBD','june')
        updateExcelOne('ManualPayments','june')
        updateExcelOne('BankAddressChange','june')
        updateExcelOne('AutoOnboarding','june')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\june.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in June,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class July(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','july')
        updateExcelOne('MBD','july')
        updateExcelOne('ManualPayments','july')
        updateExcelOne('BankAddressChange','july')
        updateExcelOne('AutoOnboarding','july')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\july.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in July,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class August(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','august')
        updateExcelOne('MBD','august')
        updateExcelOne('ManualPayments','august')
        updateExcelOne('BankAddressChange','august')
        updateExcelOne('AutoOnboarding','august')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\august.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in August,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class September(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','september')
        updateExcelOne('MBD','september')
        updateExcelOne('ManualPayments','september')
        updateExcelOne('BankAddressChange','september')
        updateExcelOne('AutoOnboarding','september')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\september.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in September,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class October(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','october')
        updateExcelOne('MBD','october')
        updateExcelOne('ManualPayments','october')
        updateExcelOne('BankAddressChange','october')
        updateExcelOne('AutoOnboarding','october')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\october.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in October,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class November(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','november')
        updateExcelOne('MBD','november')
        updateExcelOne('ManualPayments','november')
        updateExcelOne('BankAddressChange','november')
        updateExcelOne('AutoOnboarding','november')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\november.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in November,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class December(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','december')
        updateExcelOne('MBD','december')
        updateExcelOne('ManualPayments','december')
        updateExcelOne('BankAddressChange','december')
        updateExcelOne('AutoOnboarding','december')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\december.xlsx")
        sheet = wb.get_sheet_by_name("Sheet1")
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in December,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        




                    

app = SeaofBTCapp()
app.geometry('1400x850')
app.mainloop()




# In[12]:





# In[6]:





# In[ ]:




