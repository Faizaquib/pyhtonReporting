#!/usr/bin/env python
# coding: utf-8

# In[1]:


from datetime import datetime
import openpyxl
import fnmatch
import os
import tkinter as tk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import numpy as np

def countDailyFile(day,processName,Month):
    if processName == 'RTB':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\InstallsRTBMultipleTerminal\InstallsRTBMultipleTerminalFilesArchive'
    elif processName == 'MBD':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\MBD\MBDFilesArchive'
    elif processName == 'ManualPayments':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\ManualPayments\ManualPaymentsFilesArchive'
    elif processName =='AutoOnboarding':
        path = r'\\tsclient\P\Applications_XML\AutoCustomerOnboardingFilesArchive'
    elif processName == 'BankAddressChange':
        path = r'\\tsclient\P\Gateshead Ops\NS36 Workaround\CIM\BankAddress\BankAddressChangeFilesArchive'
    
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
    
    if (int(day)<10):
        day = str(0)+str(day)
     
    if (int(month)<10):
        month = str(0)+str(month)
    
    
    listOfFiles = fnmatch.filter(os.listdir(path),"*.txt")
    count = 0
    time = 0.00
    for file in listOfFiles:
        currentDate = str(day)+'-'+str(month)+'-'+str(datetime.now().year)
        fileCreationDate = datetime.date(datetime.fromtimestamp(os.stat(path+"\\"+file).st_mtime)).strftime("%d-%m-%Y")
        #iterator = iterator + 1 
        if currentDate == fileCreationDate :
             with open(path+'\\'+file) as f:
                    for line in f:
                        if 'Total Time Taken' in line:
                            count = count + 1
                            a = (line.split())
                            minutes = a[5].replace(',','')
                            time = time + float(minutes)
    return count,time


    

def updateExcelTwo(processName,Month):
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    sheet = wb["Sheet1"] 
    lastDay = str(sheet.cell(row=sheet.max_row,column=3).value)
    currentDay = str(datetime.now().day)
    currentMonth = datetime.now().month
    lastRow = sheet.max_row
    if Month == 'january':
        month = 1
    elif Month == 'feburary':
        month = 2
    elif Month == 'march':
        month = 3
    elif Month == 'april':
        month = 4
    elif Month == 'may':
        month = 5
    elif Month == 'june':
        month = 6
    elif Month == 'july':
        month = 7
    elif Month == 'august':
        month = 8
    elif Month == 'september':
        month = 9
    elif Month == 'october':
        month = 10
    elif Month == 'november':
        month = 11    
    elif Month == 'december':
        month = 12
        
        
    
    if(int(month) < int(currentMonth)):
        if(int(lastDay)<32):
            for i in range(int(lastDay)+1,32):
                lastRow = lastRow + 1
                sheet.cell(row=lastRow,column=1).value = datetime.now().year
                sheet.cell(row=lastRow,column=2).value = month
                sheet.cell(row=lastRow,column=3).value = i
                count  , time = countDailyFile(i,processName,Month)
                sheet.cell(row=lastRow,column=4).value = count
                sheet.cell(row=lastRow,column=5).value = round(time/60,2)
                wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    elif(int(month) == int(currentMonth)):
        if(int(lastDay)<int(currentDay)):
            for i in range(int(lastDay)+1,int(currentDay)+1):
                lastRow = lastRow + 1
                sheet.cell(row=lastRow,column=1).value = datetime.now().year
                sheet.cell(row=lastRow,column=2).value = month
                sheet.cell(row=lastRow,column=3).value = i
                count , time = countDailyFile(i,processName,Month)
                sheet.cell(row=lastRow,column=4).value = count
                sheet.cell(row=lastRow,column=5).value = round(time/60,2)
                wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
        elif(int(lastDay) == int(currentDay)):
            sheet.cell(row=lastRow,column=1).value = datetime.now().year
            sheet.cell(row=lastRow,column=2).value = month
            sheet.cell(row=lastRow,column=3).value = currentDay
            count , time = countDailyFile(currentDay,processName,Month)
            sheet.cell(row=lastRow,column=4).value = count
            sheet.cell(row=lastRow,column=5).value = round(time/60,2)
            wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    
    


def totalIteration(processName,Month):
    updateExcelTwo(processName,Month)
    totalIteration = 0
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    sheet = wb["Sheet1"]
    for i in range(2,sheet.max_row+1):
        totalIteration = totalIteration + sheet.cell(row=i,column=4).value
    return totalIteration

def totalTime(processName,Month):
    #pdateExcelTwo(processName,Month)
    totalTime = 0
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisTwo\\"+Month+"\\"+processName+".xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    for i in range(2,sheet.max_row+1):
        val = sheet.cell(row=i,column=5).value
        totalTime = totalTime + int(0 if val is None else val)
    return totalTime



def updateExcelOne(processName,Month):
    wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\"+Month+".xlsx")
    sheet = wb["Sheet1"]
    for i in range(2,sheet.max_row+1):
        if(sheet.cell(row=i,column=1).value == processName):
            sheet.cell(row=i,column=2).value = totalIteration(processName,Month)
            sheet.cell(row=i,column=3).value = round(totalTime(processName,Month),2)
            wb.save("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\"+Month+".xlsx")
            #print('Excel Two updated!')
            #print(Month)
                    
#updateExcelOne('RTB','december')

LARGE_FONT= ("Verdana", 12)


class SeaofBTCapp(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)

        container.pack(side="top", fill="both", expand = True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (StartPage, January,Feburary,March,April,May,June,July,August,September,October,November,December):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)

    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()

        
class StartPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        label = tk.Label(self, text="Hey There!", font=LARGE_FONT)
        label.pack(pady=50,padx=50)

        button = tk.Button(self, text="January,2020",
                            command=lambda: controller.show_frame(January))
        button.pack(side='left',pady=10,padx=10)

        button2 = tk.Button(self, text="Feburary,2020",
                            command=lambda: controller.show_frame(Feburary))
        button2.pack(side='left',pady=10,padx=10)
        
        button = tk.Button(self, text="March,2020",
                            command=lambda: controller.show_frame(March))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="April,2020",
                            command=lambda: controller.show_frame(April))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="May,2020",
                            command=lambda: controller.show_frame(May))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="June,2020",
                            command=lambda: controller.show_frame(June))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="July,2020",
                            command=lambda: controller.show_frame(July))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="August,2020",
                            command=lambda: controller.show_frame(August))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="September,2020",
                            command=lambda: controller.show_frame(September))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="October,2020",
                            command=lambda: controller.show_frame(October))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="November,2020",
                            command=lambda: controller.show_frame(November))
        button.pack(side='left',pady=10,padx=10)

        button = tk.Button(self, text="December,2020",
                            command=lambda: controller.show_frame(December))
        button.pack(side='left',pady=10,padx=10)

class January(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','january')
        updateExcelOne('MBD','january')
        updateExcelOne('ManualPayments','january')
        updateExcelOne('BankAddressChange','january')
        updateExcelOne('AutoOnboarding','january')
        print('loading modules....1/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\january.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
        timeTaken = []
        
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
            timeTaken.append(sheet.cell(row=i,column=3).value)
        xpos = np.arange(len(name))    
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in January,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            #print(name)

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        

class Feburary(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','feburary')
        updateExcelOne('MBD','feburary')
        updateExcelOne('ManualPayments','feburary')
        updateExcelOne('BankAddressChange','feburary')
        updateExcelOne('AutoOnboarding','feburary')
        print('loading modules....2/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\feburary.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in Feburary,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        

class March(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','march')
        updateExcelOne('MBD','march')
        updateExcelOne('ManualPayments','march')
        updateExcelOne('BankAddressChange','march')
        updateExcelOne('AutoOnboarding','march')
        print('loading modules....3/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\march.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in March,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)


class April(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','april')
        updateExcelOne('MBD','april')
        updateExcelOne('ManualPayments','april')
        updateExcelOne('BankAddressChange','april')
        updateExcelOne('AutoOnboarding','april')
        print('loading modules....4/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\april.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in April,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class May(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','may')
        updateExcelOne('MBD','may')
        updateExcelOne('ManualPayments','may')
        updateExcelOne('BankAddressChange','may')
        updateExcelOne('AutoOnboarding','may')
        print('loading modules....5/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\may.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in May,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class June(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','june')
        updateExcelOne('MBD','june')
        updateExcelOne('ManualPayments','june')
        updateExcelOne('BankAddressChange','june')
        updateExcelOne('AutoOnboarding','june')
        print('loading modules....6/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\june.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in June,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class July(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','july')
        updateExcelOne('MBD','july')
        updateExcelOne('ManualPayments','july')
        updateExcelOne('BankAddressChange','july')
        updateExcelOne('AutoOnboarding','july')
        print('loading modules....7/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\july.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in July,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class August(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','august')
        updateExcelOne('MBD','august')
        updateExcelOne('ManualPayments','august')
        updateExcelOne('BankAddressChange','august')
        updateExcelOne('AutoOnboarding','august')
        print('loading modules....8/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\august.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in August,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class September(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','september')
        updateExcelOne('MBD','september')
        updateExcelOne('ManualPayments','september')
        updateExcelOne('BankAddressChange','september')
        updateExcelOne('AutoOnboarding','september')
        print('loading modules....9/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\september.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in September,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class October(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','october')
        updateExcelOne('MBD','october')
        updateExcelOne('ManualPayments','october')
        updateExcelOne('BankAddressChange','october')
        updateExcelOne('AutoOnboarding','october')
        print('loading modules....10/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\october.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in October,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class November(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','november')
        updateExcelOne('MBD','november')
        updateExcelOne('ManualPayments','november')
        updateExcelOne('BankAddressChange','november')
        updateExcelOne('AutoOnboarding','november')
        print('loading modules....11/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\november.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in November,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        


class December(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        updateExcelOne('RTB','december')
        updateExcelOne('MBD','december')
        updateExcelOne('ManualPayments','december')
        updateExcelOne('BankAddressChange','december')
        updateExcelOne('AutoOnboarding','december')
        print('loading modules....12/12')
        wb = openpyxl.load_workbook("E:\\Users\\svc_genesis03\\Desktop\\genesisOne\\december.xlsx")
        sheet = wb["Sheet1"]
        name=[]
        iteration = []
       
        for i in range(2,sheet.max_row+1):
            name.append(sheet.cell(row=i,column=1).value)
            iteration.append(sheet.cell(row=i,column=2).value)
        count = 0
        for i in iteration:
            if (int(i)>0):
                count = count + 1
        if (count == 0):
            label = tk.Label(self, text="Nothing there to show!", font=LARGE_FONT)
            label.pack(pady=50,padx=50)
        else:
            f = Figure(figsize=(5,5),dpi=100)
            a = f.add_subplot(111)
            a.set_title('Genesis in December,2020')
            a.set_ylabel('Count of Iterations and Time(Hours) taken')
            a.set_xlabel('Processes Run under Genesis')
            a.set_xticks(xpos)
            a.set_xticklabels(name)
            a.bar(xpos-0.2,iteration,width =0.4,label='Iterations')
            a.bar(xpos+0.2,timeTaken,width =0.4,label='Hours')
            for i, v in enumerate(iteration):
                a.text(i-.25, v, iteration[i],fontsize=15)
            for i, v in enumerate(timeTaken):
                a.text(i+.10, v, timeTaken[i],fontsize=15)
            a.legend()
            

            canvas = FigureCanvasTkAgg(f , self)
            canvas.get_tk_widget().pack(side='bottom',fill='both', expand=True)

        button1 = tk.Button(self, text="Home",
                        command=lambda: controller.show_frame(StartPage))
        button1.pack(side='left',pady=10,padx=10)

        




                    

app = SeaofBTCapp()
app.geometry('1400x850')
app.mainloop()




# In[ ]:




